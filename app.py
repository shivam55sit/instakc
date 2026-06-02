"""
Streamlit UI for Remidio InstaKC Corneal Topography PDF extraction.

Modes:
  1. Single PDF  – upload one file, preview extracted fields, download Excel.
  2. Bulk Folder – paste a folder path, process every PDF inside, download Excel.

Run with:  streamlit run app.py
"""

import io
import tempfile
from pathlib import Path

import streamlit as st
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from extract_instakc import extract_patient_details, COLUMNS

# ── Page config ───────────────────────────────────────────────────

st.set_page_config(
    page_title="InstaKC Extractor",
    page_icon="🔬",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# ── Custom CSS ────────────────────────────────────────────────────

st.markdown(
    """
    <style>
    /* ── Google Font ─────────────────────────────────── */
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&display=swap');

    html, body, [class*="css"] {
        font-family: 'Inter', sans-serif;
    }

    /* ── Hero banner ─────────────────────────────────── */
    .hero {
        background: linear-gradient(135deg, #0f2027 0%, #203a43 40%, #2c5364 100%);
        border-radius: 16px;
        padding: 2.5rem 2rem;
        margin-bottom: 2rem;
        text-align: center;
        box-shadow: 0 8px 32px rgba(0,0,0,.35);
    }
    .hero h1 {
        color: #ffffff;
        font-size: 2.4rem;
        font-weight: 800;
        margin: 0 0 .4rem 0;
        letter-spacing: -0.5px;
    }
    .hero p {
        color: #94d2bd;
        font-size: 1.05rem;
        margin: 0;
        font-weight: 400;
    }

    /* ── Stat cards ──────────────────────────────────── */
    .stat-row {
        display: flex;
        gap: 1rem;
        margin-bottom: 1.5rem;
    }
    .stat-card {
        flex: 1;
        background: linear-gradient(135deg, #1a1a2e 0%, #16213e 100%);
        border: 1px solid rgba(148,210,189,.2);
        border-radius: 12px;
        padding: 1.2rem 1.4rem;
        text-align: center;
        box-shadow: 0 4px 16px rgba(0,0,0,.2);
    }
    .stat-card .stat-value {
        font-size: 2rem;
        font-weight: 700;
        color: #94d2bd;
    }
    .stat-card .stat-label {
        font-size: .85rem;
        color: #8899aa;
        margin-top: .25rem;
    }

    /* ── Data cards ──────────────────────────────────── */
    .data-card {
        background: linear-gradient(135deg, #1a1a2e 0%, #16213e 100%);
        border: 1px solid rgba(148,210,189,.15);
        border-radius: 14px;
        padding: 1.5rem 1.8rem;
        margin-bottom: 1rem;
        box-shadow: 0 4px 20px rgba(0,0,0,.2);
    }
    .data-card h4 {
        color: #94d2bd;
        margin: 0 0 .8rem 0;
        font-size: 1rem;
        font-weight: 600;
    }
    .field-row {
        display: flex;
        justify-content: space-between;
        padding: .45rem 0;
        border-bottom: 1px solid rgba(255,255,255,.06);
    }
    .field-row:last-child { border-bottom: none; }
    .field-label { color: #8899aa; font-size: .9rem; }
    .field-value { color: #e0e0e0; font-weight: 500; font-size: .9rem; }

    /* ── Tabs tweaks ─────────────────────────────────── */
    .stTabs [data-baseweb="tab-list"] {
        gap: 8px;
    }
    .stTabs [data-baseweb="tab"] {
        border-radius: 8px 8px 0 0;
        padding: 10px 24px;
        font-weight: 600;
    }

    /* ── Success banner ──────────────────────────────── */
    .success-banner {
        background: linear-gradient(135deg, #0d3b2e 0%, #145239 100%);
        border: 1px solid rgba(148,210,189,.3);
        border-radius: 12px;
        padding: 1rem 1.4rem;
        margin: 1rem 0;
        color: #94d2bd;
        font-weight: 500;
        text-align: center;
    }

    /* ── Hide Streamlit branding ─────────────────────── */
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
    header {visibility: hidden;}
    </style>
    """,
    unsafe_allow_html=True,
)

# ── Hero ──────────────────────────────────────────────────────────

st.markdown(
    """
    <div class="hero">
        <h1>🔬 InstaKC Extractor</h1>
        <p>Extract patient details from Remidio InstaKC Corneal Topography PDF reports</p>
    </div>
    """,
    unsafe_allow_html=True,
)


# ── Helper: build styled Excel bytes ─────────────────────────────

_HEADER_FILL = PatternFill("solid", fgColor="1F7A8C")
_HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=11)
_DATA_FONT   = Font(name="Arial", size=11)
_CENTER      = Alignment(horizontal="center", vertical="center")
_LEFT        = Alignment(horizontal="left",   vertical="center")
_THIN        = Side(style="thin", color="CCCCCC")
_BORDER      = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_COL_WIDTHS  = [22, 22, 14, 14, 22]


def _build_excel(records: list[dict]) -> bytes:
    """Create a styled Excel workbook in-memory and return raw bytes."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Patient Details"
    ws.freeze_panes = "A2"

    # Header row
    ws.row_dimensions[1].height = 22
    for ci, (col, w) in enumerate(zip(COLUMNS, _COL_WIDTHS), 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.font, cell.fill, cell.alignment, cell.border = (
            _HEADER_FONT, _HEADER_FILL, _CENTER, _BORDER,
        )
        ws.column_dimensions[cell.column_letter].width = w

    # Data rows
    for ri, rec in enumerate(records, 2):
        ws.row_dimensions[ri].height = 18
        for ci, col in enumerate(COLUMNS, 1):
            cell = ws.cell(row=ri, column=ci, value=rec.get(col, ""))
            cell.font, cell.alignment, cell.border = _DATA_FONT, _LEFT, _BORDER

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _render_detail_card(details: dict, title: str = "Extracted Patient Details"):
    """Render a single patient's details as a styled card."""
    fields_html = ""
    for k, v in details.items():
        display_val = v if v else '<span style="color:#555;">—</span>'
        fields_html += f"""
        <div class="field-row">
            <span class="field-label">{k}</span>
            <span class="field-value">{display_val}</span>
        </div>"""

    st.markdown(
        f"""<div class="data-card"><h4>{title}</h4>{fields_html}</div>""",
        unsafe_allow_html=True,
    )


def _render_stats(total: int, extracted: int, empty: int):
    """Render summary stat cards."""
    st.markdown(
        f"""
        <div class="stat-row">
            <div class="stat-card">
                <div class="stat-value">{total}</div>
                <div class="stat-label">PDFs Processed</div>
            </div>
            <div class="stat-card">
                <div class="stat-value">{extracted}</div>
                <div class="stat-label">Fields Extracted</div>
            </div>
            <div class="stat-card">
                <div class="stat-value">{empty}</div>
                <div class="stat-label">Empty Fields</div>
            </div>
        </div>""",
        unsafe_allow_html=True,
    )


# ── Tabs ──────────────────────────────────────────────────────────

tab_single, tab_bulk = st.tabs(["📄  Single PDF Upload", "📁  Bulk PDF Upload"])

# ── Tab 1 — Single Upload ─────────────────────────────────────────

with tab_single:
    st.markdown("#### Upload a single InstaKC PDF report")
    uploaded = st.file_uploader(
        "Choose a PDF file",
        type=["pdf"],
        key="single_pdf",
        help="Upload one Remidio InstaKC corneal topography report (PDF).",
    )

    if uploaded is not None:
        # Save uploaded file to a temp location for pdf2image
        with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
            tmp.write(uploaded.getvalue())
            tmp_path = tmp.name

        with st.spinner("🔍 Running OCR and extracting patient details…"):
            details = extract_patient_details(tmp_path)

        # Count stats
        total_fields = len(details)
        filled = sum(1 for v in details.values() if v)
        empty  = total_fields - filled

        _render_stats(1, filled, empty)
        _render_detail_card(details, title=f"📋  {uploaded.name}")

        # Dataframe preview
        with st.expander("📊 Table Preview", expanded=True):
            df = pd.DataFrame([details])
            st.dataframe(df, use_container_width=True, hide_index=True)

        # Download button
        excel_bytes = _build_excel([details])
        st.download_button(
            label="⬇️  Download Excel",
            data=excel_bytes,
            file_name="patient_details.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
            use_container_width=True,
        )

        st.markdown(
            '<div class="success-banner">✅ Extraction complete — download your Excel above.</div>',
            unsafe_allow_html=True,
        )

# ── Tab 2 — Bulk Folder ──────────────────────────────────────────

with tab_bulk:
    st.markdown("#### Upload multiple InstaKC PDF reports")
    bulk_files = st.file_uploader(
        "Choose PDF files",
        type=["pdf"],
        accept_multiple_files=True,
        key="bulk_pdfs",
        help="Select one or more Remidio InstaKC corneal topography reports (PDF).",
    )

    if bulk_files:
        col_start, _ = st.columns([1, 3])
        with col_start:
            start_bulk = st.button("🚀  Start Bulk Extraction", type="primary", use_container_width=True)

        if start_bulk:
            all_records: list[dict] = []
            progress_bar = st.progress(0, text="Starting…")

            for idx, pdf_file in enumerate(bulk_files):
                progress_bar.progress(
                    (idx + 1) / len(bulk_files),
                    text=f"Processing {idx + 1}/{len(bulk_files)} — {pdf_file.name}",
                )
                try:
                    # Save uploaded file to a temp location for pdf2image
                    with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
                        tmp.write(pdf_file.getvalue())
                        tmp_path = tmp.name

                    details = extract_patient_details(tmp_path)
                    details["_source_file"] = pdf_file.name
                    all_records.append(details)
                except Exception as e:
                    st.toast(f"⚠️ Skipped {pdf_file.name}: {e}", icon="⚠️")

            progress_bar.progress(1.0, text="Done ✅")

            if all_records:
                # Stats
                total_fields = sum(len(COLUMNS) for _ in all_records)
                filled = sum(
                    1
                    for rec in all_records
                    for col in COLUMNS
                    if rec.get(col)
                )
                empty = total_fields - filled

                _render_stats(len(all_records), filled, empty)

                # Show each patient card
                for i, rec in enumerate(all_records):
                    display = {k: rec.get(k, "") for k in COLUMNS}
                    _render_detail_card(
                        display,
                        title=f"📋  {rec.get('_source_file', f'Patient {i+1}')}",
                    )

                # Full dataframe
                with st.expander("📊 Full Data Table", expanded=True):
                    df = pd.DataFrame(all_records)
                    # Move source file to front
                    cols = ["_source_file"] + COLUMNS
                    df = df[[c for c in cols if c in df.columns]]
                    df = df.rename(columns={"_source_file": "Source File"})
                    st.dataframe(df, use_container_width=True, hide_index=True)

                # Build Excel (without internal _source_file column)
                excel_bytes = _build_excel(all_records)
                st.download_button(
                    label=f"⬇️  Download Excel ({len(all_records)} patients)",
                    data=excel_bytes,
                    file_name="patient_details_bulk.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary",
                    use_container_width=True,
                )

                st.markdown(
                    '<div class="success-banner">✅ Bulk extraction complete — download your Excel above.</div>',
                    unsafe_allow_html=True,
                )

# ── Footer ────────────────────────────────────────────────────────
st.markdown("---")
st.markdown(
    '<p style="text-align:center; color:#555; font-size:.82rem;">'
    "InstaKC Extractor &middot; Built with Streamlit &amp; Tesseract OCR"
    "</p>",
    unsafe_allow_html=True,
)
