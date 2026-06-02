"""
extract_instakc.py
------------------
Extracts patient details from Remidio InstaKC Corneal Topography PDF reports
and writes them to an Excel file.

HOW TO USE:
  1. Set PDF_PATH below to the path of your PDF file (or a list of PDFs).
  2. Optionally change OUTPUT_XLSX to your preferred output location.
  3. Run:  python extract_instakc.py
"""

import re
import platform
from pathlib import Path

from pdf2image import convert_from_path
import pytesseract
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# ═══════════════════════════════════════════════════════════════════
#  CONFIGURE HERE — set your PDF path(s) and output file location
# ═══════════════════════════════════════════════════════════════════

# Poppler bin directory — only needed on Windows (local dev).
# On Linux / Streamlit Cloud, poppler-utils is installed via packages.txt.
_SCRIPT_DIR = Path(__file__).resolve().parent
if platform.system() == "Windows":
    POPPLER_PATH = str(_SCRIPT_DIR / "poppler" / "poppler-24.08.0" / "Library" / "bin")
else:
    POPPLER_PATH = None  # system PATH on Linux

# Single PDF — just set the path as a string:
PDF_PATH = r"C:\Users\shivam.prajapati\Downloads\150101001006020282.pdf"

# Multiple PDFs — use a list instead (comment out the line above and
# uncomment the block below):
# PDF_PATH = [
#     r"patient1.pdf",
#     r"patient2.pdf",
#     r"patient3.pdf",
# ]

# Output Excel file — saved in the same folder as this script if left as None:
OUTPUT_XLSX = r"patient_details.xlsx"

# ═══════════════════════════════════════════════════════════════════


# ── OCR helpers ───────────────────────────────────────────────────

def _ocr_pdf(pdf_path: str, dpi: int = 200) -> str:
    """Convert PDF pages to images and run Tesseract OCR; return full text."""
    pages = convert_from_path(pdf_path, dpi=dpi, poppler_path=POPPLER_PATH)
    return "\n".join(pytesseract.image_to_string(page) for page in pages)


# ── Field extraction ──────────────────────────────────────────────

def _find(pattern: str, text: str, default: str = "") -> str:
    """Return the first capture group of a regex match, or default."""
    m = re.search(pattern, text, re.IGNORECASE)
    return m.group(1).strip() if m else default


def extract_patient_details(pdf_path: str) -> dict:
    """Extract patient header fields from a Remidio InstaKC PDF."""
    text = _ocr_pdf(pdf_path)

    name       = _find(r"Name[:\s]+([A-Za-z][^\n]{1,50}?)(?:\s{2,}|\n|MRN)", text)
    mrn        = _find(r"MRN[:\s]+([0-9]{6,})", text)
    date       = _find(r"Date[:\s]+(\d{1,2}[/\-]\d{1,2}[/\-]\d{2,4})", text)
    age_gender = _find(r"Age[/\s]*Gender[:\s]+([^\n]{2,30}?)(?:\s{2,}|\n|Doctor)", text)
    doctor     = _find(r"Doctor[\u2018\u2019's]*\s*Name[:\s]+([^\n]{2,50})", text)

    return {
        "Name":          name,
        "MRN":           mrn,
        "Date":          date,
        "Age/Gender":    age_gender,
        "Doctor's Name": doctor,
    }


# ── Excel writer ──────────────────────────────────────────────────

_HEADER_FILL = PatternFill("solid", fgColor="1F7A8C")
_HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=11)
_DATA_FONT   = Font(name="Arial", size=11)
_CENTER      = Alignment(horizontal="center", vertical="center")
_LEFT        = Alignment(horizontal="left",   vertical="center")
_THIN        = Side(style="thin", color="CCCCCC")
_BORDER      = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

COLUMNS    = ["Name", "MRN", "Date", "Age/Gender", "Doctor's Name"]
COL_WIDTHS = [22, 22, 14, 14, 22]


def _write_header(ws):
    ws.row_dimensions[1].height = 22
    for col_idx, (col_name, width) in enumerate(zip(COLUMNS, COL_WIDTHS), start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font      = _HEADER_FONT
        cell.fill      = _HEADER_FILL
        cell.alignment = _CENTER
        cell.border    = _BORDER
        ws.column_dimensions[cell.column_letter].width = width


def _write_row(ws, row_idx: int, data: dict):
    ws.row_dimensions[row_idx].height = 18
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=data.get(col_name, ""))
        cell.font      = _DATA_FONT
        cell.alignment = _LEFT
        cell.border    = _BORDER


def pdf_to_excel(pdf_path: str, output_xlsx: str) -> tuple[str, dict]:
    """
    Extract patient details from pdf_path and append a row to output_xlsx.
    Creates the file with a header row if it does not yet exist.
    Returns a tuple of (absolute path to the written Excel file, extracted details).
    """
    details = extract_patient_details(str(pdf_path))

    xlsx_path = Path(output_xlsx)
    if xlsx_path.exists():
        wb = load_workbook(output_xlsx)
        ws = wb.active
        next_row = ws.max_row + 1
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Patient Details"
        ws.freeze_panes = "A2"
        _write_header(ws)
        next_row = 2

    _write_row(ws, next_row, details)
    wb.save(output_xlsx)

    return str(xlsx_path.resolve()), details


# ── Main ──────────────────────────────────────────────────────────

if __name__ == "__main__":
    # Normalise PDF_PATH to a list so single and multi-file work the same way
    pdf_list = [PDF_PATH] if isinstance(PDF_PATH, str) else list(PDF_PATH)

    print(f"Processing {len(pdf_list)} PDF(s)...\n")

    for pdf in pdf_list:
        pdf = str(pdf)
        print(f"  → {pdf}")
        out, details = pdf_to_excel(pdf, OUTPUT_XLSX)
        for k, v in details.items():
            print(f"       {k:<18}: {v}")
        print()

    print(f"✓  Excel saved → {Path(OUTPUT_XLSX).resolve()}")